#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path

from openpyxl import load_workbook


def load_merged_rows(merged_path: Path) -> list[list[str]]:
    wb = load_workbook(merged_path)
    ws = wb.active
    rows = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if row is None:
            continue
        name = (row[0] or "").strip() if len(row) > 0 else ""
        link = (row[1] or "").strip() if len(row) > 1 else ""
        if not name and not link:
            continue
        rows.append([name, link])
    return rows


def prepend_rows_to_zsk(zsk_path: Path, rows: list[list[str]]) -> int:
    wb = load_workbook(zsk_path)
    ws = wb.active

    # 保留原表头，将新数据插在表头下面
    ws.insert_rows(2, amount=len(rows))
    for idx, row in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=row[0])
        ws.cell(row=idx, column=2, value=row[1])

    added = len(rows)
    stem = zsk_path.stem
    if stem.endswith("_zsk"):
        base = stem[: -len("_zsk")]
        output_name = f"{base}_{added}_zsk.xlsx"
    else:
        output_name = f"{stem}_{added}.xlsx"
    output_path = zsk_path.with_name(output_name)
    wb.save(output_path)
    print(f"已生成: {output_path}")
    return added


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    merged_path = base_dir / "merged_share.xlsx"
    if not merged_path.exists():
        raise SystemExit("未找到 merged_share.xlsx，请先生成合并文件")

    rows = load_merged_rows(merged_path)
    if not rows:
        raise SystemExit("merged_share.xlsx 中没有可插入的数据")

    zsk_files = sorted(base_dir.glob("*_zsk.xlsx"))
    if not zsk_files:
        raise SystemExit("未找到 *_zsk.xlsx 文件")

    for zsk_path in zsk_files:
        prepend_rows_to_zsk(zsk_path, rows)


if __name__ == "__main__":
    main()
